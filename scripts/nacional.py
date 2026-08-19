"""Unifica os Relatorio_Final_<UF>.xlsx de data/uf/ em data/nacional.json.

Cada arquivo tem o mesmo formato do relatorio de MS: aba "Relatorio", uma linha por
indicador, com pontuacao recebida e valor maximo. O script cruza tudo por codigo de
indicador para responder, em cada item, quantas UFs zeraram, quantas pontuaram em parte
e quantas atenderam integralmente - e onde MS fica nesse conjunto.

Uso:  python scripts/nacional.py [pasta]     (padrao: data/uf)
"""

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from statistics import mean, median

import openpyxl
import openpyxl.reader.excel

# Varios relatorios estaduais trazem rels apontando para um xl/drawings/drawing1.xml que
# nao existe no pacote, e o openpyxl estoura KeyError ao tentar carregar a imagem. Nada
# aqui le imagem, entao a busca por elas e neutralizada na entrada.
openpyxl.reader.excel.find_images = lambda *_a, **_k: ([], [])

RAIZ = Path(__file__).resolve().parent.parent
PADRAO = RAIZ / "data" / "uf"
SAIDA = RAIZ / "data" / "nacional.json"
FOCO = "MS"

# Colunas obrigatorias. Alguns arquivos trazem "Atendido" e "percentual" depois destas;
# nada aqui depende delas - o status e recalculado a partir de pontos e maximo.
CABECALHO = ["UF", "Dimensão", "Indicador", "Classificação recebida",
             "Pontuação recebida", "Valor do indicador"]

# Um item so responde "e problema de MS ou do pais?" com duas leituras separadas:
#   posicao de MS no item  -> MS esta atras ou na frente das outras UFs
#   fatia de UFs com nota cheia -> o item e dificil para todo mundo
# Classificar so pela fatia de zerados engana: em 2.11 nenhuma UF zerou, mas 25 das 26
# ficaram parciais - nao e fragilidade de MS, e dificuldade geral.
TERCO = 1 / 3
LIMITE_ITEM_DIFICIL = 0.35  # menos de 35% das UFs com nota cheia = obstaculo do conjunto


def classifica(pontos: float, maximo: float) -> str:
    if pontos <= 0:
        return "zerado"
    if abs(pontos - maximo) < 1e-9:
        return "cheio"
    return "parcial"


def le_uf(caminho: Path) -> tuple[str, list[dict]]:
    """Le um relatorio estadual. Devolve a sigla e as linhas normalizadas."""
    # Modo normal de proposito: varios arquivos declaram <dimension ref="A1"/>, e o modo
    # read_only confia nesse metadado e devolveria uma linha so.
    wb = openpyxl.load_workbook(caminho, data_only=True)
    ws = wb["Relatorio"] if "Relatorio" in wb.sheetnames else wb.worksheets[0]
    linhas = list(ws.iter_rows(values_only=True))

    cabecalho = [str(c).strip() if c is not None else "" for c in linhas[0]]
    if cabecalho[:len(CABECALHO)] != CABECALHO:
        raise SystemExit(
            f"{caminho.name}: cabecalho diferente do esperado.\n"
            f"  esperado: {CABECALHO}\n  encontrado: {cabecalho}"
        )

    itens = []
    for uf, dim, texto, _classificacao, pontos, maximo, *_ in linhas[1:]:
        if not uf:
            continue
        pontos, maximo = float(pontos or 0), float(maximo or 0)
        codigo = str(texto).strip().split(" ", 1)[0]
        if not re.fullmatch(r"\d+\.\d+", codigo):
            raise SystemExit(f"{caminho.name}: codigo inesperado em {texto!r}")
        itens.append({
            "uf": str(uf).strip().upper(),
            # A dimensao vem ora como 1, ora como 1.0, conforme o arquivo.
            "dimensao": str(int(float(dim))),
            "codigo": codigo,
            "pontos": pontos,
            "maximo": maximo,
            "status": classifica(pontos, maximo),
        })

    siglas = {i["uf"] for i in itens}
    if len(siglas) != 1:
        raise SystemExit(f"{caminho.name}: esperava uma UF por arquivo, achei {sorted(siglas)}")
    return siglas.pop(), itens


def posicao(valor: float, todos: list[float]) -> int:
    """Posicao em ranking decrescente, com empate recebendo a mesma colocacao."""
    return sum(1 for v in todos if v > valor + 1e-9) + 1


def main() -> None:
    pasta = Path(sys.argv[1]) if len(sys.argv) > 1 else PADRAO
    arquivos = sorted(pasta.glob("Relatorio_Final_*.xlsx"))
    if not arquivos:
        raise SystemExit(
            f"Nenhum Relatorio_Final_<UF>.xlsx em {pasta}.\n"
            f"Copie os relatorios estaduais para la e rode de novo."
        )

    por_uf: dict[str, list[dict]] = {}
    for arq in arquivos:
        sigla, itens = le_uf(arq)
        if sigla in por_uf:
            raise SystemExit(f"UF {sigla} aparece em mais de um arquivo.")
        por_uf[sigla] = itens

    if FOCO not in por_uf:
        raise SystemExit(f"O relatorio de {FOCO} precisa estar na pasta - ele e a referencia.")

    ufs = sorted(por_uf)
    qtd_ufs = len(ufs)

    # --- Ranking geral -----------------------------------------------------
    totais = {uf: round(sum(i["pontos"] for i in itens), 4) for uf, itens in por_uf.items()}
    valores = list(totais.values())
    ranking = sorted(
        ({"uf": uf, "total": t, "posicao": posicao(t, valores)} for uf, t in totais.items()),
        key=lambda r: (r["posicao"], r["uf"]),
    )

    # --- Por dimensao ------------------------------------------------------
    dimensoes = []
    for num in sorted({i["dimensao"] for i in por_uf[FOCO]}, key=int):
        aprov = {}
        for uf, itens in por_uf.items():
            do_dim = [i for i in itens if i["dimensao"] == num]
            maximo = sum(i["maximo"] for i in do_dim)
            aprov[uf] = (sum(i["pontos"] for i in do_dim) / maximo) if maximo else 0.0
        vals = list(aprov.values())
        melhor = max(aprov, key=lambda u: aprov[u])
        pior = min(aprov, key=lambda u: aprov[u])
        dimensoes.append({
            "numero": num,
            "ms": round(aprov[FOCO], 4),
            "media": round(mean(vals), 4),
            "mediana": round(median(vals), 4),
            "posicao": posicao(aprov[FOCO], vals),
            "melhor": {"uf": melhor, "valor": round(aprov[melhor], 4)},
            "pior": {"uf": pior, "valor": round(aprov[pior], 4)},
            "acima_da_media": aprov[FOCO] > mean(vals),
        })

    # --- Por indicador -----------------------------------------------------
    agrupado = defaultdict(dict)
    for uf, itens in por_uf.items():
        for i in itens:
            agrupado[i["codigo"]][uf] = i

    indicadores = []
    for i_ms in por_uf[FOCO]:
        cod = i_ms["codigo"]
        linha = agrupado[cod]
        presentes = sorted(linha)
        pontos = [linha[u]["pontos"] for u in presentes]
        zeradas = sorted(u for u in presentes if linha[u]["status"] == "zerado")
        parciais = [u for u in presentes if linha[u]["status"] == "parcial"]
        cheios = [u for u in presentes if linha[u]["status"] == "cheio"]
        share_zerado = len(zeradas) / len(presentes)
        share_cheio = len(cheios) / len(presentes)
        pos = posicao(i_ms["pontos"], pontos)

        if pos <= len(presentes) * TERCO:
            posicionamento = "frente"
        elif pos <= len(presentes) * 2 * TERCO:
            posicionamento = "meio"
        else:
            posicionamento = "atras"

        indicadores.append({
            "codigo": cod,
            "dimensao": i_ms["dimensao"],
            "ms_status": i_ms["status"],
            "ms_pontos": round(i_ms["pontos"], 4),
            "maximo": round(i_ms["maximo"], 4),
            "qtd_ufs": len(presentes),
            "zerados": len(zeradas),
            "parciais": len(parciais),
            "cheios": len(cheios),
            "share_zerado": round(share_zerado, 4),
            "share_cheio": round(share_cheio, 4),
            "media": round(mean(pontos), 4),
            "mediana": round(median(pontos), 4),
            "ms_vs_media": round(i_ms["pontos"] - mean(pontos), 4),
            "posicao": pos,
            "posicionamento": posicionamento,
            "item_dificil": share_cheio < LIMITE_ITEM_DIFICIL,
            "ufs_zeradas": zeradas,
        })

    ms_total = totais[FOCO]
    saida = {
        "uf_foco": FOCO,
        "ufs": ufs,
        "qtd_ufs": qtd_ufs,
        "ausentes": sorted(set(UF_BRASIL) - set(ufs)),
        "media_nacional": round(mean(valores), 4),
        "mediana_nacional": round(median(valores), 4),
        "melhor": ranking[0],
        "pior": ranking[-1],
        "ms": {
            "total": ms_total,
            "posicao": posicao(ms_total, valores),
            "vs_media": round(ms_total - mean(valores), 4),
            "vs_mediana": round(ms_total - median(valores), 4),
            "acima_da_media": ms_total > mean(valores),
            "acima_da_mediana": ms_total > median(valores),
        },
        # A media nacional e puxada para baixo por poucas UFs com nota muito baixa. Dizer
        # so "acima da media" mascara a posicao real, por isso a mediana anda junto.
        "abaixo_de_50": sorted(uf for uf, t in totais.items() if t < 50),
        "ranking": ranking,
        "dimensoes": dimensoes,
        "indicadores": indicadores,
        "limites": {"terco": round(TERCO, 4), "item_dificil": LIMITE_ITEM_DIFICIL},
    }

    # Self-check: o cruzamento precisa bater com o resultado de MS ja publicado.
    base = json.loads((RAIZ / "data" / "indicadores.json").read_text(encoding="utf-8"))
    assert abs(ms_total - base["resumo"]["total"]) < 0.01, (ms_total, base["resumo"]["total"])
    assert len(indicadores) == base["resumo"]["qtd"], len(indicadores)
    assert 1 <= saida["ms"]["posicao"] <= qtd_ufs
    faltando = [i["codigo"] for i in indicadores if i["qtd_ufs"] != qtd_ufs]
    if faltando:
        print(f"AVISO: itens ausentes em alguma UF: {', '.join(faltando)}")

    with SAIDA.open("w", encoding="utf-8", newline="\n") as f:
        json.dump(saida, f, ensure_ascii=False, indent=2)
        f.write("\n")
    print(
        f"OK -> {SAIDA.relative_to(RAIZ)}  ({qtd_ufs} UFs · "
        f"{FOCO} em {saida['ms']['posicao']}º com {ms_total:.2f} · "
        f"media nacional {saida['media_nacional']:.2f})"
    )
    if saida["ausentes"]:
        print(f"UFs sem relatorio na pasta: {', '.join(saida['ausentes'])}")


UF_BRASIL = [
    "AC", "AL", "AM", "AP", "BA", "CE", "DF", "ES", "GO", "MA", "MG", "MS", "MT",
    "PA", "PB", "PE", "PI", "PR", "RJ", "RN", "RO", "RR", "RS", "SC", "SE", "SP", "TO",
]

if __name__ == "__main__":
    main()
